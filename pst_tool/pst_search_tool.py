# -*- coding: utf-8 -*-
"""
Outlook PST 邮件搜索工具 v5
- 主引擎：win32com（Outlook 开着时使用，保存 .msg 格式，不乱码）
- 备用引擎：libpff（Outlook 未开时使用，保存 .eml 格式）
依赖：pip install pywin32 openpyxl libpff-python-windows
"""

import os, sys, threading, traceback, datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

try:
    import win32com.client
    import pythoncom
    HAS_WIN32 = True
except ImportError:
    HAS_WIN32 = False

try:
    import pypff
    HAS_PYPFF = True
except ImportError:
    HAS_PYPFF = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    openpyxl = None

# ── 主题标准化 ────────────────────────────────────────────────────────
_NOISE = [
    '回复:', '回复：', '回复_', '回复-', '回复 ', '回复\u3000',
    '转发:', '转发：', '转发_', '转发-', '转发 ', '转发\u3000',
    'RE:', 'RE：', 'RE_', 'RE-', 'RE ',
    'Re:', 'Re：', 'Re_', 'Re-', 'Re ',
    'FW:', 'FW：', 'FW_', 'FW-', 'FW ',
    'Fw:', 'Fw：', 'Fw_', 'Fw-', 'Fw ',
    '答复:', '答复：', '答复_', '答复-', '答复 ', '答复\u3000',
    '更新:', '更新：', '更新_', '更新-', '更新 ', '更新\u3000',
    'Recall:', 'Recall：', 'Recall_', 'Recall-', 'Recall ',
    'Update:', 'Update：', 'Update_', 'Update-', 'Update ',
    'Updated:', 'Updated：', 'Updated_', 'Updated-', 'Updated ',
    '[External]', '以此为准',
]
NOISE_PREFIXES = sorted(set(_NOISE), key=len, reverse=True)


def normalize_subject(s):
    s = (s or '').strip()
    for _ in range(20):
        changed = False
        for p in NOISE_PREFIXES:
            if len(s) >= len(p) and s[:len(p)].casefold() == p.casefold():
                s = s[len(p):].lstrip()
                changed = True
                break
        if not changed:
            break
    return s.strip()


# ── 工具函数 ──────────────────────────────────────────────────────────
def resource_path(rel):
    try:
        base = sys._MEIPASS
    except AttributeError:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, rel)


def win_long(path):
    if os.name != 'nt':
        return path
    path = os.path.abspath(path)
    if path.startswith('\\\\?\\'):
        return path
    if path.startswith('\\\\'):
        return '\\\\?\\UNC\\' + path[2:]
    return '\\\\?\\' + path


def mkdirs(path):
    os.makedirs(win_long(path), exist_ok=True)


def sanitize_filename(name, maxlen=60):
    name = (name or '无主题').strip()
    for ch in '\\/:*?"<>|\n\r\t':
        name = name.replace(ch, '_')
    return name[:maxlen].strip() or '无主题'


def safe_str(v):
    if v is None:
        return ''
    if isinstance(v, bytes):
        for enc in ('utf-8', 'gbk', 'gb2312', 'big5', 'latin-1'):
            try:
                return v.decode(enc)
            except Exception:
                pass
        return v.decode('utf-8', errors='replace')
    return str(v)


def com_date_to_datetime(com_date):
    """
    将 win32com / pywintypes.datetime 转为纯 Python 标准 datetime（无时区）。
    pywintypes.datetime 继承自 datetime.datetime 但携带时区信息，
    在某些系统/时区下 strftime 会崩溃，必须强制用属性重建为无时区对象。
    """
    if com_date is None:
        return None
    # 用 int() 强制提取各字段，兼容 pywintypes.datetime 和标准 datetime
    try:
        return datetime.datetime(
            int(com_date.year),
            int(com_date.month),
            int(com_date.day),
            int(com_date.hour),
            int(com_date.minute),
            int(com_date.second),
        )
    except Exception:
        pass
    # 兜底：转字符串再解析
    try:
        s = str(com_date)[:19]
        for fmt in ('%Y-%m-%d %H:%M:%S', '%m/%d/%Y %H:%M:%S',
                    '%Y/%m/%d %H:%M:%S'):
            try:
                return datetime.datetime.strptime(s, fmt)
            except Exception:
                pass
    except Exception:
        pass
    return None


# ── 搜索条件 ──────────────────────────────────────────────────────────
class SearchCriteria:
    def __init__(self, subject_kw='', from_kw='', to_kw='',
                 date_start=None, date_end=None, folder_filter=''):
        self.subject_kw    = subject_kw.strip().lower()
        self.from_kw       = from_kw.strip().lower()
        self.to_kw         = to_kw.strip().lower()
        self.date_start    = date_start
        self.date_end      = date_end
        self.folder_filter = folder_filter.strip()


# ── 分组管理（边搜索边写文件） ────────────────────────────────────────
class GroupManager:
    def __init__(self, eml_dir, export_eml):
        self.eml_dir     = eml_dir
        self.export_eml  = export_eml
        self.groups      = {}
        self.order       = []
        self._used_names = {}

    def _get_folder_name(self, norm_subj):
        base = sanitize_filename(norm_subj, 60)
        cnt  = self._used_names.get(base, 0)
        self._used_names[base] = cnt + 1
        return base if cnt == 0 else f'{base}_{cnt}'

    def add(self, rec_date, norm_subj, sender, save_fn):
        key = norm_subj or '(无主题)'
        if key not in self.groups:
            fname = self._get_folder_name(key)
            self.groups[key] = {
                'earliest': rec_date,
                'latest':   rec_date,
                'senders':  set(),
                'count':    0,
                'folder':   fname,
            }
            self.order.append(key)
            if self.export_eml:
                mkdirs(os.path.join(self.eml_dir, fname))
        g = self.groups[key]
        if rec_date:
            if g['earliest'] is None or rec_date < g['earliest']:
                g['earliest'] = rec_date
            if g['latest'] is None or rec_date > g['latest']:
                g['latest'] = rec_date
        if sender:
            g['senders'].add(sender)
        g['count'] += 1
        if self.export_eml and save_fn:
            folder_abs = os.path.join(self.eml_dir, g['folder'])
            try:
                save_fn(folder_abs)
            except Exception:
                pass


# ══════════════════════════════════════════════════════════════════════
# 引擎一：win32com
# ══════════════════════════════════════════════════════════════════════
def _find_pst_store(outlook, pst_path):
    norm = os.path.normpath(pst_path).lower()
    stores = outlook.Session.Stores
    for i in range(1, stores.Count + 1):
        st = stores.Item(i)
        try:
            if os.path.normpath(st.FilePath).lower() == norm:
                return st
        except Exception:
            continue
    return None


def _walk_com_folder(folder, criteria, gm, status_cb, counter):
    items = folder.Items
    items.Sort('[ReceivedTime]')

    # 用 Restrict 按时间范围预过滤，大幅减少遍历量
    filters = []
    if criteria.date_start:
        filters.append(
            f"[ReceivedTime] >= '{criteria.date_start.strftime('%Y-%m-%d')} 00:00'")
    if criteria.date_end:
        filters.append(
            f"[ReceivedTime] <= '{criteria.date_end.strftime('%Y-%m-%d')} 23:59'")
    if filters:
        items = items.Restrict(' AND '.join(filters))

    try:
        item = items.GetFirst()
        while item is not None:
            try:
                if getattr(item, 'Class', None) == 43:  # olMail
                    counter[0] += 1
                    if status_cb and counter[0] % 20 == 0:
                        status_cb(f'正在扫描第 {counter[0]} 封邮件 ...')

                    subj    = safe_str(getattr(item, 'Subject', ''))
                    sender  = safe_str(getattr(item, 'SenderName', ''))
                    to_str  = safe_str(getattr(item, 'To', ''))
                    cc_str  = safe_str(getattr(item, 'CC', ''))

                    # ── 关键修复：正确转换 COM 日期 ──
                    try:
                        raw_date = item.ReceivedTime
                        rec_date = com_date_to_datetime(raw_date)
                    except Exception:
                        rec_date = None

                    # 条件匹配
                    ok = True
                    if criteria.subject_kw:
                        ns = normalize_subject(subj).lower()
                        if criteria.subject_kw not in subj.lower() and \
                           criteria.subject_kw not in ns:
                            ok = False
                    if ok and criteria.from_kw:
                        se = safe_str(getattr(item, 'SenderEmailAddress', ''))
                        if criteria.from_kw not in sender.lower() and \
                           criteria.from_kw not in se.lower():
                            ok = False
                    if ok and criteria.to_kw:
                        pool = (to_str + ' ' + cc_str).lower()
                        if criteria.to_kw not in pool:
                            ok = False

                    if ok:
                        norm_subj = normalize_subject(subj)
                        ts = rec_date.strftime('%Y%m%d_%H%M%S') \
                            if rec_date else '无日期'
                        fname = f"{ts}_{sanitize_filename(norm_subj, 40)}.msg"
                        captured = item

                        def save_fn(folder_abs,
                                    _it=captured, _fn=fname):
                            fp = os.path.join(folder_abs, _fn)
                            if not os.path.exists(fp):
                                _it.SaveAs(win_long(fp))

                        gm.add(rec_date, norm_subj, sender, save_fn)

            except Exception:
                pass
            try:
                item = items.GetNext()
            except Exception:
                break
    except Exception:
        pass

    # 递归子文件夹
    try:
        for i in range(1, folder.Folders.Count + 1):
            sub = folder.Folders.Item(i)
            sub_name = safe_str(getattr(sub, 'Name', ''))
            if criteria.folder_filter:
                if sub_name == criteria.folder_filter:
                    _walk_com_folder(sub, criteria, gm, status_cb, counter)
                else:
                    # 继续向深处找目标文件夹
                    _walk_com_folder_find(sub, criteria, gm, status_cb, counter)
            else:
                _walk_com_folder(sub, criteria, gm, status_cb, counter)
    except Exception:
        pass


def _walk_com_folder_find(folder, criteria, gm, status_cb, counter):
    """只找名称匹配的子文件夹，不扫描当前层邮件"""
    try:
        for i in range(1, folder.Folders.Count + 1):
            sub = folder.Folders.Item(i)
            sub_name = safe_str(getattr(sub, 'Name', ''))
            if sub_name == criteria.folder_filter:
                _walk_com_folder(sub, criteria, gm, status_cb, counter)
            else:
                _walk_com_folder_find(sub, criteria, gm, status_cb, counter)
    except Exception:
        pass


def run_search_com(pst_path, criteria, gm, status_cb=None):
    pythoncom.CoInitialize()
    try:
        outlook = win32com.client.Dispatch('Outlook.Application')
        store   = _find_pst_store(outlook, pst_path)
        if store is None:
            outlook.Session.AddStore(pst_path)
            store = _find_pst_store(outlook, pst_path)
        if store is None:
            raise RuntimeError(
                f'无法在 Outlook 中找到该 PST：{pst_path}\n'
                '请先在 Outlook 里手动打开这个 PST 文件。')

        root    = store.GetRootFolder()
        counter = [0]

        if criteria.folder_filter:
            _walk_com_folder_find(root, criteria, gm, status_cb, counter)
        else:
            _walk_com_folder(root, criteria, gm, status_cb, counter)

        return counter[0]
    finally:
        pythoncom.CoUninitialize()


# ══════════════════════════════════════════════════════════════════════
# 引擎二：libpff（备用）
# ══════════════════════════════════════════════════════════════════════
def run_search_pff(pst_path, criteria, gm, status_cb=None):
    pst_path = os.path.normpath(os.path.abspath(pst_path))
    pf = pypff.file()
    pf.open(pst_path)
    root    = pf.get_root_folder()
    counter = [0]

    def get_date(msg):
        for attr in ('delivery_time', 'client_submit_time', 'creation_time'):
            try:
                v = getattr(msg, attr, None)
                if v:
                    return v
            except Exception:
                pass
        return None

    def walk(folder):
        for msg in folder.sub_messages:
            counter[0] += 1
            if status_cb and counter[0] % 25 == 0:
                status_cb(f'正在扫描第 {counter[0]} 封邮件 ...')
            try:
                subj    = safe_str(getattr(msg, 'subject', ''))
                sender  = safe_str(getattr(msg, 'sender_name', ''))
                to_str  = safe_str(getattr(msg, 'display_to', ''))
                cc_str  = safe_str(getattr(msg, 'display_cc', ''))
                headers = safe_str(getattr(msg, 'transport_headers', ''))
                rec_date = get_date(msg)

                # 时间范围快速跳过
                if (criteria.date_start or criteria.date_end) and rec_date:
                    d = rec_date.date() if hasattr(rec_date, 'date') else None
                    if d:
                        if criteria.date_start and d < criteria.date_start:
                            continue
                        if criteria.date_end and d > criteria.date_end:
                            continue

                ok = True
                ns = normalize_subject(subj).lower()
                if criteria.subject_kw and \
                   criteria.subject_kw not in subj.lower() and \
                   criteria.subject_kw not in ns:
                    ok = False
                if ok and criteria.from_kw:
                    if criteria.from_kw not in sender.lower() and \
                       criteria.from_kw not in headers.lower():
                        ok = False
                if ok and criteria.to_kw:
                    pool = (to_str + ' ' + cc_str + ' ' + headers).lower()
                    if criteria.to_kw not in pool:
                        ok = False

                if ok:
                    norm_subj = normalize_subject(subj)
                    ts = rec_date.strftime('%Y%m%d_%H%M%S') \
                        if rec_date else '无日期'
                    fname = f"{ts}_{sanitize_filename(norm_subj, 40)}.eml"
                    plain = safe_str(getattr(msg, 'plain_text_body', ''))
                    html  = safe_str(getattr(msg, 'html_body', ''))

                    def save_fn(folder_abs,
                                _s=subj, _sndr=sender, _to=to_str,
                                _cc=cc_str, _dt=rec_date,
                                _p=plain, _h=html, _fn=fname):
                        import quopri
                        lines = ['MIME-Version: 1.0',
                                 f'Subject: {_s}',
                                 f'From: {_sndr}',
                                 f'To: {_to}']
                        if _cc:
                            lines.append(f'Cc: {_cc}')
                        if _dt:
                            try:
                                from email.utils import format_datetime
                                lines.append(f'Date: {format_datetime(_dt)}')
                            except Exception:
                                pass
                        if _h:
                            lines += ['Content-Type: text/html; charset="utf-8"',
                                      'Content-Transfer-Encoding: quoted-printable', '',
                                      quopri.encodestring(_h.encode('utf-8')).decode('ascii')]
                        else:
                            lines += ['Content-Type: text/plain; charset="utf-8"',
                                      'Content-Transfer-Encoding: quoted-printable', '',
                                      quopri.encodestring(
                                          (_p or '(无正文)').encode('utf-8')).decode('ascii')]
                        with open(win_long(os.path.join(folder_abs, _fn)), 'wb') as f:
                            f.write('\r\n'.join(lines).encode('utf-8'))

                    gm.add(rec_date, norm_subj, sender, save_fn)
            except Exception:
                pass

        for sub in folder.sub_folders:
            name = safe_str(getattr(sub, 'name', '')) or '未命名'
            if criteria.folder_filter:
                if name == criteria.folder_filter:
                    walk(sub)
                else:
                    walk_find(sub)
            else:
                walk(sub)

    def walk_find(folder):
        for sub in folder.sub_folders:
            name = safe_str(getattr(sub, 'name', '')) or '未命名'
            if name == criteria.folder_filter:
                walk(sub)
            else:
                walk_find(sub)

    try:
        if criteria.folder_filter:
            walk_find(root)
        else:
            walk(root)
    finally:
        pf.close()

    return counter[0]


# ── 主搜索入口 ────────────────────────────────────────────────────────
def run_search(pst_path, criteria, gm, status_cb=None):
    if HAS_WIN32:
        status_cb and status_cb('使用 Outlook COM 引擎（保存为 .msg 格式）...')
        return run_search_com(pst_path, criteria, gm, status_cb)
    elif HAS_PYPFF:
        status_cb and status_cb('使用 libpff 引擎（保存为 .eml 格式）...')
        return run_search_pff(pst_path, criteria, gm, status_cb)
    else:
        raise RuntimeError('未安装搜索引擎，请执行：pip install pywin32 openpyxl')


# ── 生成 Excel 报表 ───────────────────────────────────────────────────
def export_excel(gm, total, output_dir, status_cb=None):
    if openpyxl is None:
        raise RuntimeError('未安装 openpyxl，请执行：pip install openpyxl')

    wb = openpyxl.Workbook()

    # ── Sheet1：检索主题邮件（5列）──────────────────────────────────
    ws = wb.active
    ws.title = '检索主题邮件'

    # 表头样式
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4472C4')
    header_align = Alignment(horizontal='center')

    headers = ['序号', '主题', '接收时间', '最新邮件时间', '邮件发件人']
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    # 数据行
    for g_idx, key in enumerate(gm.order, 1):
        if status_cb and g_idx % 50 == 0:
            status_cb(f'正在写入报表第 {g_idx} 行...')
        g = gm.groups[key]

        # 日期格式化（已是标准 datetime，直接格式化）
        earliest_str = g['earliest'].strftime('%Y-%m-%d %H:%M') \
            if g['earliest'] else ''
        latest_str = g['latest'].strftime('%Y-%m-%d %H:%M') \
            if g['latest'] else ''
        senders_str = '; '.join(sorted(g['senders']))

        ws.cell(row=g_idx + 1, column=1, value=g_idx)
        ws.cell(row=g_idx + 1, column=2, value=key)
        ws.cell(row=g_idx + 1, column=3, value=earliest_str)
        ws.cell(row=g_idx + 1, column=4, value=latest_str)
        ws.cell(row=g_idx + 1, column=5, value=senders_str)

        # 主题列加超链接
        if gm.export_eml:
            folder_abs = os.path.abspath(
                os.path.join(gm.eml_dir, g['folder']))
            link_url = 'file:///' + folder_abs.replace('\\', '/')
            c = ws.cell(row=g_idx + 1, column=2)
            c.hyperlink = link_url
            c.font = Font(color='0563C1', underline='single')

    for col, w in zip('ABCDE', [6, 46, 18, 18, 34]):
        ws.column_dimensions[col].width = w

    # ── Sheet2：统计汇总 ─────────────────────────────────────────────
    ws2 = wb.create_sheet('统计汇总')
    total_matched = sum(g['count'] for g in gm.groups.values())
    rate = (total_matched / total * 100) if total else 0
    rows = [
        ('生成时间',  datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ('PST邮件总数', total),
        ('匹配邮件数',  total_matched),
        ('匹配率',      f'{rate:.2f}%'),
        ('识别主题数',  len(gm.order)),
        ('搜索引擎',    'Outlook COM (.msg)' if HAS_WIN32 else 'libpff (.eml)'),
    ]
    for r in rows:
        ws2.append(r)
    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 28

    report = os.path.join(
        output_dir,
        f"邮件搜索报表_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb.save(win_long(report))
    return report, rate, total_matched


# ── 界面 ──────────────────────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title('Outlook PST 邮件搜索工具')
        self.geometry('680x700')
        self.resizable(False, False)

        self.pst_path_var   = tk.StringVar()
        self.folder_var     = tk.StringVar()
        self.output_dir_var = tk.StringVar()
        self.subject_var    = tk.StringVar()
        self.from_var       = tk.StringVar()
        self.to_var         = tk.StringVar()
        self.date_start_var = tk.StringVar()
        self.date_end_var   = tk.StringVar()
        self.export_eml_var = tk.BooleanVar(value=True)
        self.status_var     = tk.StringVar(value='就绪')

        self.logo_img = None
        try:
            self.logo_img = tk.PhotoImage(
                file=resource_path(os.path.join('assets', 'logo.png')))
        except Exception:
            pass

        self._build_ui()

    def _build_ui(self):
        pad = {'padx': 10, 'pady': 5}

        # ── 版权（先 pack，确保始终显示在底部）──
        ttk.Label(self,
                  text='Copyright © 2026 CTCI Beijing Co., Ltd.',
                  foreground='#999999',
                  font=('Arial', 8)).pack(side='bottom', pady=(2, 6))

        # ── 顶部 logo ──
        frm_hdr = ttk.Frame(self)
        frm_hdr.pack(fill='x', padx=14, pady=(10, 4))
        if self.logo_img:
            ttk.Label(frm_hdr, image=self.logo_img).pack(side='left')
        ttk.Label(frm_hdr, text='Outlook PST 邮件搜索工具',
                  font=('Microsoft YaHei UI', 11, 'bold'),
                  foreground='#3c3c3c').pack(side='right', padx=4)

        # ── 第一步 ──
        frm_file = ttk.LabelFrame(self, text='第一步：选择 PST 文件与输出文件夹')
        frm_file.pack(fill='x', **pad)

        ttk.Label(frm_file, text='PST 文件：').grid(
            row=0, column=0, sticky='w', padx=6, pady=5)
        ttk.Entry(frm_file, textvariable=self.pst_path_var, width=52).grid(
            row=0, column=1, pady=5)
        ttk.Button(frm_file, text='浏览...', command=self.choose_pst).grid(
            row=0, column=2, padx=6)

        ttk.Label(frm_file, text='指定子文件夹：').grid(
            row=1, column=0, sticky='w', padx=6, pady=5)
        ttk.Entry(frm_file, textvariable=self.folder_var, width=52).grid(
            row=1, column=1, pady=5, sticky='w')
        ttk.Label(frm_file,
                  text='可选，填写 PST 内子文件夹名，留空则搜全部',
                  foreground='#666666').grid(
            row=2, column=0, columnspan=3, sticky='w', padx=6)

        ttk.Label(frm_file, text='结果输出到：').grid(
            row=3, column=0, sticky='w', padx=6, pady=5)
        ttk.Entry(frm_file, textvariable=self.output_dir_var, width=52).grid(
            row=3, column=1, pady=5)
        ttk.Button(frm_file, text='浏览...', command=self.choose_output).grid(
            row=3, column=2, padx=6)

        ttk.Label(frm_file,
                  text='提示：若 PST 正被 Outlook 打开，请先关闭 Outlook 或使用副本。',
                  foreground='#a15c00').grid(
            row=4, column=0, columnspan=3, sticky='w', padx=6, pady=(0, 5))

        # ── 第二步 ──
        frm_s = ttk.LabelFrame(self, text='第二步：设置搜索条件（可任意组合，留空不限制）')
        frm_s.pack(fill='x', **pad)

        ttk.Label(frm_s, text='主题关键字：').grid(
            row=0, column=0, sticky='w', padx=6, pady=5)
        ttk.Entry(frm_s, textvariable=self.subject_var, width=38).grid(
            row=0, column=1, columnspan=2, sticky='w', pady=5)

        ttk.Label(frm_s, text='发件人包含：').grid(
            row=1, column=0, sticky='w', padx=6, pady=5)
        ttk.Entry(frm_s, textvariable=self.from_var, width=38).grid(
            row=1, column=1, columnspan=2, sticky='w', pady=5)

        ttk.Label(frm_s, text='收件人/抄送：').grid(
            row=2, column=0, sticky='w', padx=6, pady=5)
        ttk.Entry(frm_s, textvariable=self.to_var, width=38).grid(
            row=2, column=1, columnspan=2, sticky='w', pady=5)

        ttk.Label(frm_s, text='日期范围：').grid(
            row=3, column=0, sticky='w', padx=6, pady=5)
        ttk.Entry(frm_s, textvariable=self.date_start_var, width=14).grid(
            row=3, column=1, sticky='w', pady=5)
        ttk.Label(frm_s, text='至').grid(
            row=3, column=1, padx=(112, 0), sticky='w')
        ttk.Entry(frm_s, textvariable=self.date_end_var, width=14).grid(
            row=3, column=2, sticky='w', pady=5)
        ttk.Label(frm_s, text='格式：YYYY-MM-DD（可只填一端）',
                  foreground='#666666').grid(
            row=4, column=0, columnspan=3, sticky='w', padx=6)

        ttk.Checkbutton(
            frm_s,
            text='将匹配邮件另存为文件（按主题分子文件夹，Outlook 开启时保存为 .msg）',
            variable=self.export_eml_var).grid(
            row=5, column=0, columnspan=3, sticky='w', padx=6, pady=(6, 4))

        # ── 搜索按钮 ──
        self.search_btn = ttk.Button(self, text='开始搜索并生成报表',
                                     command=self.start_search)
        self.search_btn.pack(pady=8)

        # ── 进度区域（最后 pack，expand 占满剩余空间）──
        frm_prog = ttk.LabelFrame(self, text='进度与结果')
        frm_prog.pack(fill='both', expand=True, **pad)

        self.progress = ttk.Progressbar(frm_prog, mode='indeterminate')
        self.progress.pack(fill='x', padx=10, pady=8)

        ttk.Label(frm_prog, textvariable=self.status_var,
                  wraplength=600, justify='left').pack(
            anchor='w', padx=10, pady=2)

        self.result_text = tk.Text(frm_prog, height=7, wrap='word')
        self.result_text.pack(fill='both', expand=True, padx=10, pady=6)
        self.result_text.configure(state='disabled')

    # ── 交互 ──────────────────────────────────────────────────────────
    def choose_pst(self):
        path = filedialog.askopenfilename(
            title='选择 PST 文件',
            filetypes=[('Outlook 数据文件', '*.pst'), ('所有文件', '*.*')])
        if path:
            path = os.path.normpath(path)
            self.pst_path_var.set(path)
            if not self.output_dir_var.get():
                self.output_dir_var.set(
                    os.path.join(os.path.dirname(path), '搜索结果'))

    def choose_output(self):
        path = filedialog.askdirectory(title='选择结果输出文件夹')
        if path:
            self.output_dir_var.set(os.path.normpath(path))

    def parse_date(self, s):
        s = s.strip()
        if not s:
            return None
        try:
            return datetime.datetime.strptime(s, '%Y-%m-%d').date()
        except ValueError:
            raise ValueError(f'日期格式不正确：{s}，请用 YYYY-MM-DD')

    def set_status(self, text):
        self.status_var.set(text)
        self.update_idletasks()

    def append_result(self, text):
        self.result_text.configure(state='normal')
        self.result_text.insert('end', text + '\n')
        self.result_text.configure(state='disabled')
        self.result_text.see('end')

    def start_search(self):
        pst_path   = self.pst_path_var.get().strip()
        output_dir = self.output_dir_var.get().strip()

        if not pst_path or not os.path.isfile(pst_path):
            messagebox.showerror('错误', '请先选择有效的 PST 文件')
            return
        if not output_dir:
            messagebox.showerror('错误', '请选择结果输出文件夹')
            return
        if not HAS_WIN32 and not HAS_PYPFF:
            messagebox.showerror('缺少依赖',
                '请执行：pip install pywin32 openpyxl libpff-python-windows')
            return

        try:
            date_start = self.parse_date(self.date_start_var.get())
            date_end   = self.parse_date(self.date_end_var.get())
        except ValueError as e:
            messagebox.showerror('日期格式错误', str(e))
            return

        criteria = SearchCriteria(
            subject_kw    = self.subject_var.get(),
            from_kw       = self.from_var.get(),
            to_kw         = self.to_var.get(),
            date_start    = date_start,
            date_end      = date_end,
            folder_filter = self.folder_var.get(),
        )

        self.search_btn.configure(state='disabled')
        self.progress.start(12)
        self.result_text.configure(state='normal')
        self.result_text.delete('1.0', 'end')
        self.result_text.configure(state='disabled')

        threading.Thread(
            target=self._run,
            args=(pst_path, output_dir, criteria, self.export_eml_var.get()),
            daemon=True,
        ).start()

    def _run(self, pst_path, output_dir, criteria, export_eml):
        try:
            mkdirs(output_dir)
            eml_dir = os.path.join(output_dir, '匹配邮件')
            if export_eml:
                mkdirs(eml_dir)

            gm = GroupManager(eml_dir, export_eml)
            self.set_status('正在搜索，匹配到的邮件将实时保存...')
            total = run_search(pst_path, criteria, gm, self.set_status)

            self.set_status('搜索完成，正在生成 Excel 报表...')
            report, rate, matched = export_excel(
                gm, total, output_dir, self.set_status)

            self.set_status('完成！')
            engine = 'Outlook COM (.msg)' if HAS_WIN32 else 'libpff (.eml)'
            self.append_result(f'搜索引擎：{engine}')
            self.append_result(f'PST 文件：{pst_path}')
            self.append_result(
                f'扫描总数：{total}　匹配：{matched}　匹配率：{rate:.2f}%')
            self.append_result(f'识别主题（线程）数：{len(gm.order)}')
            self.append_result(f'报表：{report}')
            if export_eml:
                self.append_result(f'邮件导出到：{eml_dir}')

            messagebox.showinfo('搜索完成',
                f'扫描 {total} 封，匹配 {matched} 封，匹配率 {rate:.2f}%\n'
                f'识别主题数：{len(gm.order)}\n报表保存到：\n{output_dir}')
        except Exception as e:
            traceback.print_exc()
            self.set_status('出现错误')
            messagebox.showerror('执行出错', str(e))
        finally:
            self.progress.stop()
            self.search_btn.configure(state='normal')


def main():
    app = App()
    app.mainloop()


if __name__ == '__main__':
    main()
